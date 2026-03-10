"""
Airzone Linky Energy Integration
==================================
Fetches electricity consumption from the Conso API (Enedis/Linky smart meter),
stores 30-minute interval readings, and analyzes heat pump consumption by
correlating on/off states with total energy usage.

No UI dependencies — usable by both macOS app and Pi daemon.

Setup:
    1. Visit https://conso.boris.sh/ and authenticate with your Enedis account
    2. Copy the Bearer token and your 14-digit PRM (meter ID)
    3. Add to config: linky_enabled, linky_token, linky_prm
"""
from __future__ import annotations

import logging
import sqlite3
import statistics
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import requests
except ImportError:
    requests = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

log = logging.getLogger("airzone")

CONSO_API_BASE = "https://conso.boris.sh/api"


# ── DB Schema ────────────────────────────────────────────────────────────────

def create_linky_tables(conn: sqlite3.Connection):
    """Create Linky energy tables if they don't exist."""
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS linky_readings (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT    NOT NULL,
            wh        REAL    NOT NULL,
            source    TEXT    DEFAULT 'load_curve',
            UNIQUE(timestamp)
        );

        CREATE INDEX IF NOT EXISTS idx_linky_ts
            ON linky_readings(timestamp);

        CREATE TABLE IF NOT EXISTS energy_analysis (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            date                TEXT    NOT NULL,
            total_kwh           REAL,
            base_kwh            REAL,
            heatpump_kwh        REAL,
            heating_hours       REAL,
            avg_outdoor_temp    REAL,
            kwh_per_heating_hr  REAL,
            outdoor_temp_band   TEXT,
            UNIQUE(date)
        );

        CREATE INDEX IF NOT EXISTS idx_energy_date
            ON energy_analysis(date);
    """)
    conn.commit()


# ── Conso API Client ────────────────────────────────────────────────────────

def _api_get(endpoint: str, token: str, prm: str,
             start: date, end: date) -> dict:
    """Make a GET request to the Conso API."""
    if requests is None:
        raise ImportError("requests package is required for Linky integration")

    headers = {
        "Authorization": f"Bearer {token}",
        "User-Agent": "airzone-daemon/1.0",
    }
    params = {
        "prm": prm,
        "start": str(start),
        "end": str(end),
    }
    resp = requests.get(
        f"{CONSO_API_BASE}/{endpoint}",
        headers=headers,
        params=params,
        timeout=30,
    )

    if resp.status_code == 401:
        log.error("Linky: token expired or invalid (401). "
                  "Re-authenticate at https://conso.boris.sh/")
        return None
    if resp.status_code == 429:
        log.warning("Linky: rate limited (429), will retry next cycle")
        return None
    if resp.status_code != 200:
        log.error("Linky API error %d: %s", resp.status_code, resp.text[:200])
        return None

    return resp.json()


def fetch_load_curve(token: str, prm: str,
                     start: date, end: date) -> list[dict]:
    """
    Fetch 30-minute consumption intervals.

    Returns list of {timestamp: str, wh: float} dicts.
    The API returns W averaged over 30min; we convert to Wh (value / 2).
    Max 7 days per request.
    """
    # Clamp to max 7 days
    if (end - start).days > 7:
        end = start + timedelta(days=7)

    data = _api_get("consumption_load_curve", token, prm, start, end)
    if not data:
        return []

    readings = []
    try:
        # API format changed: interval_reading may be at top level
        # or nested under meter_reading
        if "meter_reading" in data:
            intervals = data["meter_reading"]["interval_reading"]
        elif "interval_reading" in data:
            intervals = data["interval_reading"]
        else:
            log.error("Linky: unexpected response format: %s",
                      list(data.keys()))
            return []
        for iv in intervals:
            watts_avg = float(iv["value"])
            wh = watts_avg / 2  # W averaged over 30min → Wh
            readings.append({
                "timestamp": iv["date"],
                "wh": round(wh, 1),
            })
    except (KeyError, TypeError, ValueError) as e:
        log.error("Linky: failed to parse load curve: %s", e)
        return []

    log.info("Linky: fetched %d load curve intervals (%s to %s)",
             len(readings), start, end)
    return readings


def fetch_daily_consumption(token: str, prm: str,
                            start: date, end: date) -> list[dict]:
    """
    Fetch daily consumption totals.

    Returns list of {date: str, wh: float} dicts.
    """
    data = _api_get("daily_consumption", token, prm, start, end)
    if not data:
        return []

    readings = []
    try:
        if "meter_reading" in data:
            intervals = data["meter_reading"]["interval_reading"]
        elif "interval_reading" in data:
            intervals = data["interval_reading"]
        else:
            log.error("Linky: unexpected daily response format: %s",
                      list(data.keys()))
            return []
        for iv in intervals:
            readings.append({
                "date": iv["date"][:10],
                "wh": float(iv["value"]),
            })
    except (KeyError, TypeError, ValueError) as e:
        log.error("Linky: failed to parse daily consumption: %s", e)
        return []

    log.info("Linky: fetched %d daily readings (%s to %s)",
             len(readings), start, end)
    return readings


# ── Storage ──────────────────────────────────────────────────────────────────

def store_load_curve(conn: sqlite3.Connection, readings: list[dict]):
    """Bulk insert load curve readings (ignores duplicates)."""
    if not readings:
        return 0
    conn.executemany(
        "INSERT OR IGNORE INTO linky_readings (timestamp, wh, source) "
        "VALUES (?, ?, 'load_curve')",
        [(r["timestamp"], r["wh"]) for r in readings],
    )
    conn.commit()
    return len(readings)


# ── Heat Pump Inference ──────────────────────────────────────────────────────

def _get_heating_state_for_slots(conn: sqlite3.Connection,
                                 day: date) -> dict[str, bool]:
    """
    For each 30-min slot on a given day, determine if any zone had heating ON.

    Returns dict mapping slot start timestamp (HH:MM) to bool.
    Uses zone_readings power field — if ANY reading in a slot has power=1,
    the slot is considered "heating".
    """
    day_str = str(day)
    next_day_str = str(day + timedelta(days=1))

    rows = conn.execute(
        "SELECT timestamp, power FROM zone_readings "
        "WHERE timestamp >= ? AND timestamp < ? "
        "ORDER BY timestamp",
        (day_str, next_day_str)
    ).fetchall()

    # Build 48 slots (00:00, 00:30, 01:00, ...)
    slots = {}
    for h in range(24):
        for m in (0, 30):
            slot_key = f"{h:02d}:{m:02d}"
            slots[slot_key] = False

    for row in rows:
        ts_str = row[0] if isinstance(row[0], str) else str(row[0])
        power = row[1]
        if power:
            # Find which 30-min slot this reading belongs to
            try:
                if "T" in ts_str:
                    time_part = ts_str.split("T")[1][:5]
                elif " " in ts_str:
                    time_part = ts_str.split(" ")[1][:5]
                else:
                    continue
                hour = int(time_part[:2])
                minute = int(time_part[3:5])
                slot_minute = 0 if minute < 30 else 30
                slot_key = f"{hour:02d}:{slot_minute:02d}"
                slots[slot_key] = True
            except (ValueError, IndexError):
                continue

    return slots


def _get_outdoor_temps_for_day(conn: sqlite3.Connection,
                               day: date) -> list[float]:
    """Get all outdoor temperature readings for a day."""
    day_str = str(day)
    next_day_str = str(day + timedelta(days=1))
    rows = conn.execute(
        "SELECT outdoor_temp FROM zone_readings "
        "WHERE timestamp >= ? AND timestamp < ? "
        "AND outdoor_temp IS NOT NULL",
        (day_str, next_day_str)
    ).fetchall()
    return [r[0] for r in rows]


def _temp_band(temp: float) -> str:
    """Round temperature to 5-degree band label."""
    lower = int(temp // 5) * 5
    upper = lower + 5
    return f"{lower}-{upper}\u00b0C"


def analyze_energy(conn: sqlite3.Connection, day: date) -> dict | None:
    """
    Analyze energy consumption for a single day.

    Correlates Linky 30-min readings with zone heating states to infer
    base load vs heat pump consumption.
    """
    day_str = str(day)
    next_day_str = str(day + timedelta(days=1))

    # Get Linky readings for this day
    rows = conn.execute(
        "SELECT timestamp, wh FROM linky_readings "
        "WHERE timestamp >= ? AND timestamp < ? "
        "ORDER BY timestamp",
        (day_str, next_day_str)
    ).fetchall()

    if len(rows) < 10:  # Need reasonable coverage
        return None

    linky_by_slot = {}
    for ts, wh in rows:
        ts_str = ts if isinstance(ts, str) else str(ts)
        try:
            if " " in ts_str:
                time_part = ts_str.split(" ")[1][:5]
            elif "T" in ts_str:
                time_part = ts_str.split("T")[1][:5]
            else:
                continue
            # Conso API timestamps mark the END of the interval
            # So "00:30:00" = the 00:00-00:30 slot
            hour = int(time_part[:2])
            minute = int(time_part[3:5])
            if minute == 0 and hour > 0:
                # e.g. "01:00" → slot "00:30"
                slot_key = f"{hour - 1:02d}:30"
            elif minute == 0 and hour == 0:
                # "00:00" → previous day's last slot, skip
                continue
            elif minute == 30:
                slot_key = f"{hour:02d}:00"
            else:
                slot_key = f"{hour:02d}:{0 if minute < 30 else 30:02d}"
            linky_by_slot[slot_key] = wh
        except (ValueError, IndexError):
            continue

    if not linky_by_slot:
        return None

    # Get heating states
    heating_slots = _get_heating_state_for_slots(conn, day)

    # Split into idle and heating
    idle_wh = []
    heating_wh = []
    total_wh = 0

    for slot_key, wh in linky_by_slot.items():
        total_wh += wh
        is_heating = heating_slots.get(slot_key, False)
        if is_heating:
            heating_wh.append(wh)
        else:
            idle_wh.append(wh)

    # Base consumption = median of idle slots
    if idle_wh:
        base_per_slot = statistics.median(idle_wh)
    else:
        # All slots had heating — use minimum as rough base estimate
        base_per_slot = min(linky_by_slot.values()) if linky_by_slot else 0

    # Heat pump consumption
    heatpump_total_wh = sum(max(0, wh - base_per_slot) for wh in heating_wh)
    heating_hours = len(heating_wh) * 0.5  # Each slot = 30 min
    base_total_wh = base_per_slot * len(linky_by_slot)

    # Outdoor temperature
    outdoor_temps = _get_outdoor_temps_for_day(conn, day)
    avg_outdoor = (sum(outdoor_temps) / len(outdoor_temps)
                   if outdoor_temps else None)

    kwh_per_hr = (heatpump_total_wh / 1000 / heating_hours
                  if heating_hours > 0 else None)

    return {
        "date": day_str,
        "total_kwh": round(total_wh / 1000, 2),
        "base_kwh": round(base_total_wh / 1000, 2),
        "heatpump_kwh": round(heatpump_total_wh / 1000, 2),
        "heating_hours": round(heating_hours, 1),
        "avg_outdoor_temp": round(avg_outdoor, 1) if avg_outdoor else None,
        "kwh_per_heating_hr": round(kwh_per_hr, 2) if kwh_per_hr else None,
        "outdoor_temp_band": (_temp_band(avg_outdoor)
                              if avg_outdoor is not None else None),
    }


# ── Full Analysis Orchestrator ───────────────────────────────────────────────

def run_energy_analysis(conn: sqlite3.Connection,
                        days: int = 30) -> dict:
    """
    Analyze energy for all available days and store results.

    Returns summary with daily stats and temperature band efficiency.
    """
    create_linky_tables(conn)

    cutoff = date.today() - timedelta(days=days)
    # Find days with Linky data
    rows = conn.execute(
        "SELECT DISTINCT substr(timestamp, 1, 10) AS day "
        "FROM linky_readings WHERE timestamp >= ? "
        "ORDER BY day",
        (str(cutoff),)
    ).fetchall()

    analyzed = 0
    for (day_str,) in rows:
        try:
            day = date.fromisoformat(day_str)
        except ValueError:
            continue

        result = analyze_energy(conn, day)
        if not result:
            continue

        conn.execute(
            "INSERT OR REPLACE INTO energy_analysis "
            "(date, total_kwh, base_kwh, heatpump_kwh, heating_hours, "
            " avg_outdoor_temp, kwh_per_heating_hr, outdoor_temp_band) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (result["date"], result["total_kwh"], result["base_kwh"],
             result["heatpump_kwh"], result["heating_hours"],
             result["avg_outdoor_temp"], result["kwh_per_heating_hr"],
             result["outdoor_temp_band"]),
        )
        analyzed += 1

    conn.commit()
    log.info("Linky: analyzed energy for %d/%d days", analyzed, len(rows))

    # Build summary
    return _build_energy_summary(conn, days)


def _build_energy_summary(conn: sqlite3.Connection,
                          days: int = 30) -> dict:
    """Build energy analysis summary from stored results."""
    cutoff = str(date.today() - timedelta(days=days))

    # Daily stats
    daily = conn.execute(
        "SELECT date, total_kwh, base_kwh, heatpump_kwh, heating_hours, "
        "       avg_outdoor_temp, kwh_per_heating_hr, outdoor_temp_band "
        "FROM energy_analysis WHERE date >= ? ORDER BY date DESC",
        (cutoff,)
    ).fetchall()

    daily_stats = [dict(zip(
        ["date", "total_kwh", "base_kwh", "heatpump_kwh", "heating_hours",
         "avg_outdoor_temp", "kwh_per_heating_hr", "outdoor_temp_band"], r))
        for r in daily]

    # Temperature band efficiency
    bands = conn.execute(
        "SELECT outdoor_temp_band, "
        "       AVG(kwh_per_heating_hr) AS avg_kwh_hr, "
        "       SUM(heating_hours) AS total_hours, "
        "       COUNT(*) AS days_count "
        "FROM energy_analysis "
        "WHERE date >= ? AND kwh_per_heating_hr IS NOT NULL "
        "      AND outdoor_temp_band IS NOT NULL "
        "GROUP BY outdoor_temp_band "
        "ORDER BY outdoor_temp_band",
        (cutoff,)
    ).fetchall()

    temp_bands = [dict(zip(
        ["temp_band", "avg_kwh_per_hr", "total_heating_hours", "days"], r))
        for r in bands]
    for b in temp_bands:
        b["avg_kwh_per_hr"] = round(b["avg_kwh_per_hr"], 2)
        b["total_heating_hours"] = round(b["total_heating_hours"], 1)

    # Savings estimate
    savings = compute_savings(conn, days)

    return {
        "daily_stats": daily_stats,
        "temp_band_efficiency": temp_bands,
        "savings": savings,
        "analyzed_at": datetime.utcnow().isoformat() + "Z",
    }


def compute_savings(conn: sqlite3.Connection, days: int = 30) -> dict | None:
    """
    Compare heat pump efficiency during warm window hours vs cold hours.

    Uses heating_cycles start times cross-referenced with energy_analysis
    and outdoor temperatures to estimate savings from COP optimization.
    """
    cutoff = str(date.today() - timedelta(days=days))

    rows = conn.execute(
        "SELECT date, kwh_per_heating_hr, avg_outdoor_temp "
        "FROM energy_analysis "
        "WHERE date >= ? AND kwh_per_heating_hr IS NOT NULL "
        "      AND avg_outdoor_temp IS NOT NULL",
        (cutoff,)
    ).fetchall()

    if len(rows) < 5:
        return None

    # Split into warm days (outdoor > 10°C) and cold days (outdoor <= 10°C)
    warm_rates = [r[1] for r in rows if r[2] > 10]
    cold_rates = [r[1] for r in rows if r[2] <= 10]

    if not warm_rates or not cold_rates:
        return None

    avg_warm = statistics.mean(warm_rates)
    avg_cold = statistics.mean(cold_rates)

    if avg_cold == 0:
        return None

    savings_pct = round((1 - avg_warm / avg_cold) * 100, 1) if avg_cold > avg_warm else 0

    return {
        "warm_kwh_per_hr": round(avg_warm, 2),
        "cold_kwh_per_hr": round(avg_cold, 2),
        "savings_pct": savings_pct,
        "warm_days": len(warm_rates),
        "cold_days": len(cold_rates),
        "reasoning": (
            f"Heating when outdoor >10\u00b0C uses {avg_warm:.2f} kWh/h "
            f"vs {avg_cold:.2f} kWh/h when \u226410\u00b0C "
            f"({savings_pct}% less energy)"
        ),
    }


# ── File Import (Enedis Excel/CSV) ───────────────────────────────────────────

def import_enedis_file(conn: sqlite3.Connection, filepath: str) -> dict:
    """
    Import Enedis 'courbe de charge' export (Excel or CSV) into linky_readings.

    Enedis exports have:
      - Excel: sheet 'Consommation Horaire', data starts at row 17
      - Columns: [_, _, Start datetime, End datetime, Value (kW), _]
      - Datetime format: DD/MM/YYYY HH:MM:SS
      - Value is average power in kW over the 30-min slot
      - We store Wh = kW × 1000 × 0.5 (30 min)

    Returns {imported: int, skipped: int, start_date: str, end_date: str}
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    rows_data = []

    if path.suffix.lower() in (".xlsx", ".xls"):
        if openpyxl is None:
            raise ImportError(
                "openpyxl is required for Excel import: "
                "pip install openpyxl")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        # Try known sheet name, fall back to first sheet
        if "Consommation Horaire" in wb.sheetnames:
            ws = wb["Consommation Horaire"]
        else:
            ws = wb.active

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < 17:  # Skip header rows
                continue
            # Columns: [None, None, Start, End, Value_kW, None]
            if len(row) < 5 or row[2] is None or row[4] is None:
                continue
            rows_data.append((row[2], row[4]))

    elif path.suffix.lower() == ".csv":
        import csv
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=";")
            for i, row in enumerate(reader, start=1):
                if i < 17 or len(row) < 5:
                    continue
                start_str = row[2].strip() if row[2] else None
                val_str = row[4].strip() if row[4] else None
                if not start_str or not val_str:
                    continue
                try:
                    val = float(val_str.replace(",", "."))
                except ValueError:
                    continue
                rows_data.append((start_str, val))
    else:
        raise ValueError(
            f"Unsupported file format: {path.suffix}. "
            "Use .xlsx or .csv from Enedis export.")

    if not rows_data:
        return {"imported": 0, "skipped": 0,
                "start_date": None, "end_date": None}

    # Parse and insert
    create_linky_tables(conn)
    imported = 0
    skipped = 0
    timestamps = []

    for start_raw, value_kw in rows_data:
        try:
            # Parse datetime — could be string or datetime object from openpyxl
            if isinstance(start_raw, datetime):
                dt = start_raw
            elif isinstance(start_raw, str):
                # Try DD/MM/YYYY HH:MM:SS format (Enedis standard)
                for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                            "%Y-%m-%dT%H:%M:%S"):
                    try:
                        dt = datetime.strptime(start_raw, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    skipped += 1
                    continue
            else:
                skipped += 1
                continue

            # Convert kW average over 30 min to Wh
            kw = float(value_kw)
            wh = round(kw * 1000 * 0.5, 1)  # kW → W × 0.5h = Wh

            # Store as ISO format timestamp (matches Conso API format)
            ts_iso = dt.strftime("%Y-%m-%d %H:%M:%S")
            timestamps.append(dt)

            conn.execute(
                "INSERT OR IGNORE INTO linky_readings "
                "(timestamp, wh, source) VALUES (?, ?, 'enedis_import')",
                (ts_iso, wh),
            )
            imported += 1
        except (ValueError, TypeError):
            skipped += 1
            continue

    conn.commit()

    start_date = min(timestamps).strftime("%Y-%m-%d") if timestamps else None
    end_date = max(timestamps).strftime("%Y-%m-%d") if timestamps else None

    log.info("Linky import: %d readings imported, %d skipped (%s to %s)",
             imported, skipped, start_date, end_date)

    return {
        "imported": imported,
        "skipped": skipped,
        "start_date": start_date,
        "end_date": end_date,
    }


# ── Convenience ──────────────────────────────────────────────────────────────

def get_energy_readings(conn: sqlite3.Connection,
                        hours: int = 168) -> list[dict]:
    """Get raw Linky readings for chart display."""
    cutoff = (datetime.utcnow() - timedelta(hours=hours)).isoformat() + "Z"
    rows = conn.execute(
        "SELECT timestamp, wh FROM linky_readings "
        "WHERE timestamp >= ? ORDER BY timestamp",
        (cutoff,)
    ).fetchall()
    return [{"timestamp": r[0], "wh": r[1]} for r in rows]


def get_energy_analysis(conn: sqlite3.Connection,
                        days: int = 30) -> list[dict]:
    """Get daily energy analysis for display."""
    cutoff = str(date.today() - timedelta(days=days))
    rows = conn.execute(
        "SELECT date, total_kwh, base_kwh, heatpump_kwh, heating_hours, "
        "       avg_outdoor_temp, kwh_per_heating_hr, outdoor_temp_band "
        "FROM energy_analysis WHERE date >= ? ORDER BY date DESC",
        (cutoff,)
    ).fetchall()
    return [dict(zip(
        ["date", "total_kwh", "base_kwh", "heatpump_kwh", "heating_hours",
         "avg_outdoor_temp", "kwh_per_heating_hr", "outdoor_temp_band"], r))
        for r in rows]


def get_temp_band_efficiency(conn: sqlite3.Connection,
                             days: int = 30) -> list[dict]:
    """Get kWh/h grouped by outdoor temperature band."""
    cutoff = str(date.today() - timedelta(days=days))
    rows = conn.execute(
        "SELECT outdoor_temp_band, "
        "       ROUND(AVG(kwh_per_heating_hr), 2) AS avg_kwh_hr, "
        "       ROUND(SUM(heating_hours), 1) AS total_hours, "
        "       COUNT(*) AS days_count "
        "FROM energy_analysis "
        "WHERE date >= ? AND kwh_per_heating_hr IS NOT NULL "
        "      AND outdoor_temp_band IS NOT NULL "
        "GROUP BY outdoor_temp_band "
        "ORDER BY outdoor_temp_band",
        (cutoff,)
    ).fetchall()
    return [dict(zip(
        ["temp_band", "avg_kwh_per_hr", "total_heating_hours", "days"], r))
        for r in rows]
